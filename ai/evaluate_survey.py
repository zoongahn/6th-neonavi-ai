"""층2 검증 — 설문 응답과 모델 추천의 인간 일치율.

층1(train.py의 test 정확도)은 **규칙을 학습했다**만 증명한다. 라벨이 규칙에서
나왔으니 순환이다(R1). 이 모듈은 그 순환 밖에서, 실제 사람 48명이 고른 경로와
모델의 랭킹이 얼마나 맞는지를 잰다.

설문 경로는 우리가 수집한 경로가 아니라 **설문용으로 서술된 가상 경로**다.
따라서 학습 데이터 누수는 없다(격리 불필요).

⚠️ 표현 정합 한계 — 설문은 12특성 중 일부만 제공한다:
   제공: duration_min, toll, curvature, slope, signal_count, road_type
   없음: distance_km, avg_speed, congestion, turn_count, fuel_cost, speed_limit
   없는 특성은 후보 전체에 같은 값을 넣어 vectorize.normalize 가 0.5(중립)로
   만들게 둔다. 지어내지 않는다. 대신 fuel 축은 3특성 중 2개가 빠져 약해진다.

⚠️ 프로필도 4필드뿐 — passenger·load_kg 는 설문에서 묻지 않았다. 모두 같은 값
   (alone/0)으로 두므로 이 두 입력은 변별에 기여하지 않는다.

실행: .venv/bin/python -m ai.evaluate_survey
"""
import collections
import os

import openpyxl

from .features import vectorize
from .recommender import baseline_b, model_a
from .recommender import weights as rule_weights
from .schema import PREFERENCE_AXES

_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
_DEFAULT_XLSX = os.path.join(_DATA_DIR, 'survey',
                             'NeoNavi_설문결과_모델학습용_정제본.xlsx')

SHEET = '학습용_후보별'

# 설문 컬럼 → 우리 특성. 방향(높을수록 나쁨/좋음)이 계약과 같은 것만 싣는다.
DIRECT = {
    'duration_min': 'duration_min',
    'toll_won': 'toll',
    'curvature_score': 'curvature',
    'slope_score': 'slope',
    'signal_score': 'signal_count',
    'road_score': 'road_type',
}


def load_rows(path=None) -> list:
    path = path or _DEFAULT_XLSX
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    return [dict(zip(header, r)) for r in it if r and r[0]]


def to_route(row) -> dict:
    """설문 후보 1개 → 특성 dict. 없는 특성은 넣지 않는다(후보 전체에 없으면 중립)."""
    route = {'id': row['candidate_route_id']}
    for src, dst in DIRECT.items():
        value = row.get(src)
        if value is not None:
            route[dst] = float(value)
    return route


def to_profile(row) -> dict:
    """설문 프로필 4필드 → 모델 입력. passenger·load_kg 는 설문에 없어 고정값.

    설문 코드북은 SUV 를 대문자로 남긴다. 계약(encoders.CAR_TYPES)은 소문자라
    그대로 넣으면 one-hot 이 전부 0 이 되어 차종 정보가 조용히 사라진다.
    """
    car_age = str(row.get('car_age') or '')
    years = 2.0 if '≤3' in car_age else (5.5 if '4' in car_age else 9.0)
    car_type = str(row.get('car_type') or 'sedan').strip().lower()
    return {
        'age': int(row['age']),
        'gender': row['gender'],
        'car_type': car_type,
        'car_age': years,
        'passenger': 'alone',   # 설문 미수집
        'load_kg': 0.0,         # 설문 미수집
    }


def group_by_response(rows) -> list:
    """(respondent, scenario) 단위로 후보 3개 + 사람이 고른 것."""
    groups = collections.OrderedDict()
    for row in rows:
        key = (row['respondent_id'], row['scenario_id'])
        groups.setdefault(key, []).append(row)

    out = []
    for (rid, sid), items in groups.items():
        chosen = [i['candidate_route_id'] for i in items if int(i['selected'] or 0) == 1]
        if len(chosen) != 1:
            continue                     # 선택이 없거나 중복이면 채점 불가
        out.append({
            'respondent_id': rid,
            'scenario_id': sid,
            'profile': to_profile(items[0]),
            'routes': [to_route(i) for i in items],
            'human_choice': chosen[0],
        })
    return out


def _top1(recs) -> str:
    return recs[0].route_id if recs else ''


def _rank_map(recs) -> dict:
    """route_id → 순위(0이 1순위)."""
    return {r.route_id: i for i, r in enumerate(recs)}


def majority_ceiling(cases) -> float:
    """개인화를 전혀 안 할 때의 상한.

    시나리오마다 '가장 많이 고른 경로' 하나만 내놓는 모델의 정확도. 프로필을 보지
    않는 어떤 모델도 이보다 잘할 수 없다. 개인화의 가치는 이 선을 넘는 만큼이다.
    """
    by_scenario = collections.defaultdict(collections.Counter)
    for c in cases:
        by_scenario[c['scenario_id']][c['human_choice']] += 1
    hit = sum(counter.most_common(1)[0][1] for counter in by_scenario.values())
    return hit / len(cases) if cases else 0.0


def evaluate(path=None, verbose=True) -> dict:
    """모델·규칙의 top-1 이 사람 선택과 일치하는 비율."""
    cases = group_by_response(load_rows(path))
    handle = model_a.load_model()

    hit_model = hit_rule = 0
    by_scenario = collections.defaultdict(lambda: [0, 0, 0])   # [모델, 규칙, 총]
    human_dist = collections.Counter()
    model_dist = collections.Counter()

    pair_model = pair_rule = pair_total = 0

    for case in cases:
        human = case['human_choice']
        m_recs = model_a.recommend(case['profile'], case['routes'], model=handle)
        r_recs = baseline_b.recommend(case['profile'], case['routes'])
        m_top, r_top = _top1(m_recs), _top1(r_recs)

        hit_model += m_top == human
        hit_rule += r_top == human
        s = by_scenario[case['scenario_id']]
        s[0] += m_top == human
        s[1] += r_top == human
        s[2] += 1
        human_dist[human] += 1
        model_dist[m_top] += 1

        # pairwise: 사람이 고른 것 vs 나머지 각각. 층1 정확도와 같은 단위라 비교 가능.
        m_rank, r_rank = _rank_map(m_recs), _rank_map(r_recs)
        for route in case['routes']:
            other = route['id']
            if other == human:
                continue
            pair_total += 1
            pair_model += m_rank.get(human, 99) < m_rank.get(other, 99)
            pair_rule += r_rank.get(human, 99) < r_rank.get(other, 99)

    n = len(cases)
    result = {
        'n_cases': n,
        'n_respondents': len({c['respondent_id'] for c in cases}),
        'model_top1': hit_model / n if n else 0.0,
        'rule_top1': hit_rule / n if n else 0.0,
        'chance': 1 / 3,
        'majority_ceiling': majority_ceiling(cases),
        'model_pairwise': pair_model / pair_total if pair_total else 0.0,
        'rule_pairwise': pair_rule / pair_total if pair_total else 0.0,
        'n_pairs': pair_total,
        'by_scenario': {k: {'model': v[0] / v[2], 'rule': v[1] / v[2], 'n': v[2]}
                        for k, v in by_scenario.items()},
    }

    if verbose:
        print(f'응답자 {result["n_respondents"]}명 × 시나리오 → 채점 가능 {n}건\n')
        print(f'  사람 선택과 top-1 일치율')
        print(f'    모델(model_a) : {result["model_top1"]*100:5.1f}%')
        print(f'    규칙(baseline): {result["rule_top1"]*100:5.1f}%')
        print(f'    무작위 기준선  : {result["chance"]*100:5.1f}%')
        print(f'    개인화 무시 상한: {result["majority_ceiling"]*100:5.1f}%'
              '  ← 시나리오별 최빈 선택만 내놓는 모델')
        print(f'\n  pairwise 일치율 (n={pair_total}) — 층1 정확도와 같은 단위')
        print(f'    모델 {result["model_pairwise"]*100:5.1f}%   '
              f'규칙 {result["rule_pairwise"]*100:5.1f}%   무작위 50.0%\n')
        print('  시나리오별')
        for sid, v in sorted(result['by_scenario'].items()):
            print(f'    {sid}  모델 {v["model"]*100:5.1f}%  규칙 {v["rule"]*100:5.1f}%  (n={v["n"]})')
        print('\n  선택 분포 (사람 vs 모델)')
        for rid in sorted(set(human_dist) | set(model_dist)):
            print(f'    {rid:6s} 사람 {human_dist[rid]:3d}   모델 {model_dist[rid]:3d}')
    return result


def _axis_matrix(case):
    """후보별 성향축 만족도 + id 목록."""
    norms = vectorize.normalize(
        [vectorize.build_feature_vector(r) for r in case['routes']])
    return [r['id'] for r in case['routes']], [vectorize.project_to_axes(n) for n in norms]


def _fit_weights(rows, steps=600, lr=0.05):
    """사람이 고른 경로가 나머지보다 높게 나오도록 축 가중치를 적합(pairwise logistic)."""
    import torch

    diffs = []
    for ids, ax, human in rows:
        hv = torch.tensor([ax[ids.index(human)][a] for a in PREFERENCE_AXES])
        for i, rid in enumerate(ids):
            if rid != human:
                diffs.append(hv - torch.tensor([ax[i][a] for a in PREFERENCE_AXES]))
    if not diffs:
        return None
    x = torch.stack(diffs)
    w = torch.zeros(len(PREFERENCE_AXES), requires_grad=True)
    opt = torch.optim.Adam([w], lr=lr)
    for _ in range(steps):
        opt.zero_grad()
        torch.nn.functional.softplus(-(x @ w)).mean().backward()
        opt.step()
    return w.detach()


def refit(path=None, group_fn=None, folds=3, verbose=True) -> dict:
    """축 가중치를 사람 데이터로 적합해 held-out 평가.

    group_fn 을 주면 그 그룹별로 따로 적합한다(개인화 검증용). 표본이 15건 미만인
    그룹은 전역 가중치로 대신한다 — 적은 표본에 맞추면 과적합만 는다.
    """
    import torch

    cases = group_by_response(load_rows(path))
    data = []
    for c in cases:
        ids, ax = _axis_matrix(c)
        data.append({'r': c['respondent_id'], 'ids': ids, 'ax': ax,
                     'human': c['human_choice'], 'case': c})
    rids = sorted({d['r'] for d in data})
    key = group_fn or (lambda case: 'ALL')

    top1 = pair = ptot = n = 0
    for k in range(folds):
        held = {r for i, r in enumerate(rids) if i % folds == k}
        tr = [d for d in data if d['r'] not in held]
        glob = _fit_weights([(d['ids'], d['ax'], d['human']) for d in tr])
        ws = {}
        for g in {key(d['case']) for d in tr}:
            sub = [d for d in tr if key(d['case']) == g]
            ws[g] = (_fit_weights([(d['ids'], d['ax'], d['human']) for d in sub])
                     if len(sub) >= 15 else glob)

        for d in (x for x in data if x['r'] in held):
            w = ws.get(key(d['case']), glob)
            scores = [float(torch.tensor([d['ax'][i][a] for a in PREFERENCE_AXES]) @ w)
                      for i in range(len(d['ids']))]
            top1 += d['ids'][max(range(len(scores)), key=lambda i: scores[i])] == d['human']
            hi = d['ids'].index(d['human'])
            for i in range(len(scores)):
                if i != hi:
                    ptot += 1
                    pair += scores[hi] > scores[i]
            n += 1

    whole = _fit_weights([(d['ids'], d['ax'], d['human']) for d in data])
    pos = whole.clamp(min=0)
    result = {
        'top1': top1 / n, 'pairwise': pair / ptot,
        'weights_raw': {a: round(float(whole[i]), 2) for i, a in enumerate(PREFERENCE_AXES)},
        'weights_norm': {a: round(float((pos / pos.sum())[i]), 3)
                         for i, a in enumerate(PREFERENCE_AXES)},
    }
    if verbose:
        print(f'  top-1 {result["top1"]*100:5.1f}%   pairwise {result["pairwise"]*100:5.1f}%'
              f'   가중치 {result["weights_norm"]}')
    return result


def main():
    evaluate()
    print('\n── 축 가중치를 사람 데이터로 재적합 (응답자 held-out) ──')
    print('  개인화 없음:', end=' ')
    refit()
    print('  규칙 성향별:', end=' ')
    refit(group_fn=lambda c: rule_weights.profile_to_mode(c['profile']))


if __name__ == '__main__':
    main()
